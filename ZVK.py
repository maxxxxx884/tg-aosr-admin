import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import subprocess
import platform
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class IncomingJournalEditor:
    def __init__(self, root):
        self.root = root
        self.root.title("Редактор журналов входного контроля")
        self.root.geometry("1700x700")  # Увеличил ширину для дополнительного столбца

        self.data = []
        self.current_directory = ""

        # Для inline редактирования
        self.edit_item = None
        self.edit_column = None
        self.entry_widget = None

        # Для сортировки
        self.sort_column = None
        self.sort_reverse = False

        self.create_widgets()

    def create_widgets(self):
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Button(top_frame, text="Выбрать директорию",
                   command=self.select_directory).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Обновить данные",
                   command=self.load_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Сохранить изменения",
                   command=self.save_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Экспорт в Excel",
                   command=self.export_to_excel).pack(side=tk.LEFT, padx=5)

        self.info_label = ttk.Label(top_frame, text="Выберите директорию для начала работы")
        self.info_label.pack(side=tk.LEFT, padx=20)

        # Инструкция по использованию
        instruction_label = ttk.Label(top_frame,
                                      text="Двойной клик - редактирование | Клик по заголовку - сортировка | Ctrl+Click - открыть документы")
        instruction_label.pack(side=tk.RIGHT, padx=10)

        self.create_treeview()

    def create_treeview(self):
        tree_frame = ttk.Frame(self.root)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Столбцы таблицы для журнала входного контроля (добавлен "Проверка док.")
        self.columns = ("Подрядчик", "Дата", "Наименование материала", "Количество", "Ед. изм.",
                        "Поставщик", "Документ", "Проверка кач.", "Файлы", "Лаб. контроль", "Результат лаб.",
                        "Исполнитель", "Создано")

        # Соответствие столбцов ключам в данных
        self.column_keys = {
            "Подрядчик": "contractor",
            "Дата": "date",
            "Наименование материала": "name",
            "Количество": "quantity",
            "Ед. изм.": "quantity_unit",
            "Поставщик": "supplier",
            "Документ": "document",
            "Проверка кач.": "document_check_result",  # Добавлено новое поле
            "Файлы": "document_files",
            "Лаб. контроль": "lab_control_needed",
            "Результат лаб.": "lab_control_result",
            "Исполнитель": "filled_by",
            "Создано": "created_at"
        }

        # Редактируемые столбцы (добавлено "Проверка док.")
        self.editable_columns = {"Дата", "Наименование материала", "Количество", "Ед. изм.",
                                 "Поставщик", "Документ", "Проверка кач.", "Лаб. контроль", "Результат лаб.",
                                 "Исполнитель"}

        self.tree = ttk.Treeview(tree_frame, columns=self.columns, show='headings', height=20)

        # Настройка заголовков и ширины столбцов (добавлена ширина для нового столбца)
        column_widths = [100, 80, 180, 80, 60, 120, 100, 100, 60, 80, 100, 120, 120]
        for i, col in enumerate(self.columns):
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_by_column(c))
            self.tree.column(col, width=column_widths[i])

        # Скроллбары
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Размещение элементов
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Привязка событий
        self.tree.bind('<Double-1>', self.on_double_click)
        self.tree.bind('<Control-1>', self.on_ctrl_click)
        self.tree.bind('<Button-1>', self.on_single_click)

    def select_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.current_directory = directory
            self.info_label.config(text=f"Директория: {directory}")
            self.load_data()

    def get_contractor_name(self, path):
        """Определение имени подрядчика из пути"""
        path_parts = os.path.normpath(path).split(os.sep)
        if "Журнал входного контроля" in path_parts:
            idx = path_parts.index("Журнал входного контроля")
            if idx > 0:
                return path_parts[idx - 1]
        return os.path.basename(os.path.dirname(path)) or "Неизвестный подрядчик"

    def find_contractor_root(self, contractor_name):
        """Поиск корневой папки подрядчика в основной директории"""
        if not self.current_directory:
            return None

        for root, dirs, files in os.walk(self.current_directory):
            if os.path.basename(root) == contractor_name:
                return root
        return None

    def load_data(self):
        if not self.current_directory:
            messagebox.showwarning("Предупреждение", "Сначала выберите директорию")
            return

        self.data = []

        # Поиск всех файлов journal_incoming.json
        for root, dirs, files in os.walk(self.current_directory):
            if "journal_incoming.json" in files:
                contractor = self.get_contractor_name(root)
                file_path = os.path.join(root, "journal_incoming.json")

                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        journal_data = json.load(f)

                    for entry in journal_data.get('entries', []):
                        entry['contractor'] = contractor
                        entry['file_path'] = file_path
                        entry['root_path'] = root
                        entry['contractor_root'] = self.find_contractor_root(contractor)
                        self.data.append(entry)

                except Exception as e:
                    messagebox.showerror("Ошибка", f"Ошибка при загрузке {file_path}: {str(e)}")

        self.update_treeview()
        self.info_label.config(text=f"Загружено записей: {len(self.data)}")

    def update_treeview(self):
        # Сохраняем текущий выбор
        selected = self.tree.selection()

        # Очищаем таблицу
        self.tree.delete(*self.tree.get_children())

        # Заполняем таблицы данными
        for i, entry in enumerate(self.data):
            # Формируем отображение количества
            quantity_display = ""
            if isinstance(entry.get('quantity'), (int, float)):
                quantity_display = str(entry.get('quantity', ''))
            else:
                quantity_display = entry.get('quantity', '')

            # Количество файлов документов
            doc_files = entry.get('document_files', [])
            files_text = f"{len(doc_files)} файл(ов)" if doc_files else "Нет"

            # Лабораторный контроль
            lab_control = "Да" if entry.get('lab_control_needed', False) else "Нет"

            # Добавляем значение проверки документов
            doc_check = entry.get('document_check_result', '')

            values = (
                entry.get('contractor', ''),
                entry.get('date', ''),
                entry.get('name', ''),
                quantity_display,
                entry.get('quantity_unit', ''),
                entry.get('supplier', ''),
                entry.get('document', ''),
                doc_check,  # Новый столбец
                files_text,
                lab_control,
                entry.get('lab_control_result', ''),
                entry.get('filled_by', ''),
                entry.get('created_at', '')
            )
            self.tree.insert('', 'end', iid=str(i), values=values)

        # Восстанавливаем выбор если возможно
        if selected and selected[0] in [self.tree.get_children()[i] for i in range(len(self.data))]:
            self.tree.selection_set(selected[0])

    def sort_by_column(self, column):
        """Сортировка по столбцу"""
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_reverse = False

        self.sort_column = column

        # Получаем ключ для сортировки
        sort_key = self.column_keys.get(column, column.lower())

        def sort_function(entry):
            value = entry.get(sort_key, '')
            if sort_key == 'quantity':
                try:
                    if isinstance(value, (int, float)):
                        return float(value)
                    return float(value) if value else 0
                except (ValueError, TypeError):
                    return 0
            elif sort_key == 'document_files':
                return len(entry.get('document_files', []))
            elif sort_key == 'lab_control_needed':
                return entry.get('lab_control_needed', False)
            return str(value).lower()

        # Сортируем данные
        self.data.sort(key=sort_function, reverse=self.sort_reverse)

        # Обновляем отображение
        self.update_treeview()

        # Обновляем заголовок с индикатором сортировки
        for col in self.columns:
            if col == column:
                direction = " ↓" if self.sort_reverse else " ↑"
                self.tree.heading(col, text=col + direction)
            else:
                self.tree.heading(col, text=col)

    def on_single_click(self, event):
        """Обработка одиночного клика - завершение редактирования"""
        self.finish_edit()

    def on_double_click(self, event):
        """Обработка двойного клика - начало редактирования"""
        self.finish_edit()  # Завершаем предыдущее редактирование

        item = self.tree.selection()[0] if self.tree.selection() else None
        if not item:
            return

        column = self.tree.identify_column(event.x)
        if not column:
            return

        column_index = int(column.replace('#', '')) - 1
        column_name = self.columns[column_index]

        # Проверяем, можно ли редактировать этот столбец
        if column_name not in self.editable_columns:
            return

        self.start_edit(item, column_name, column_index)

    def on_ctrl_click(self, event):
        """Обработка Ctrl+Click - открытие документов"""
        item = self.tree.selection()[0] if self.tree.selection() else None
        if item:
            self.open_documents_for_item(item)

    def start_edit(self, item, column_name, column_index):
        """Начало редактирования ячейки"""
        self.edit_item = item
        self.edit_column = column_name

        # Получаем координаты и размеры ячейки
        bbox = self.tree.bbox(item, column_index)
        if not bbox:
            return

        x, y, width, height = bbox

        # Получаем текущее значение
        current_value = self.tree.item(item, 'values')[column_index]

        # Для поля "Лаб. контроль" создаем комбобокс
        if column_name == "Лаб. контроль":
            self.entry_widget = ttk.Combobox(self.tree, values=["Да", "Нет"], state="readonly")
            self.entry_widget.place(x=x, y=y, width=width, height=height)
            self.entry_widget.set(current_value)
        # Для поля "Проверка док." создаем комбобокс с типичными значениями
        elif column_name == "Проверка кач.":
            self.entry_widget = ttk.Combobox(self.tree,
                                             values=["соответствует", "не соответствует", "требует доработки",
                                                     "на доработке"])
            self.entry_widget.place(x=x, y=y, width=width, height=height)
            self.entry_widget.set(current_value)
        else:
            # Создаем Entry для редактирования
            self.entry_widget = tk.Entry(self.tree)
            self.entry_widget.place(x=x, y=y, width=width, height=height)
            self.entry_widget.insert(0, current_value)
            self.entry_widget.select_range(0, tk.END)

        self.entry_widget.focus()

        # Привязываем события
        self.entry_widget.bind('<Return>', self.finish_edit)
        self.entry_widget.bind('<Escape>', self.cancel_edit)
        self.entry_widget.bind('<FocusOut>', self.finish_edit)

    def finish_edit(self, event=None):
        """Завершение редактирования"""
        if not self.entry_widget or not self.edit_item:
            return

        # Получаем новое значение
        new_value = self.entry_widget.get()

        # Обновляем данные
        item_index = int(self.edit_item)
        column_key = self.column_keys[self.edit_column]

        # Специальная обработка для разных типов полей
        if column_key == 'quantity':
            try:
                new_value = float(new_value) if new_value.strip() else None
            except ValueError:
                messagebox.showerror("Ошибка", "Количество должно быть числом")
                return
        elif column_key == 'lab_control_needed':
            new_value = new_value == "Да"

        # Обновляем данные
        self.data[item_index][column_key] = new_value

        # Обновляем время изменения
        self.data[item_index]['created_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Убираем Entry
        self.entry_widget.destroy()
        self.entry_widget = None
        self.edit_item = None
        self.edit_column = None

        # Обновляем отображение
        self.update_treeview()

    def cancel_edit(self, event=None):
        """Отмена редактирования"""
        if self.entry_widget:
            self.entry_widget.destroy()
            self.entry_widget = None
            self.edit_item = None
            self.edit_column = None

    def open_documents_for_item(self, item):
        """Открытие документов для конкретного элемента"""
        item_index = int(item)
        entry = self.data[item_index]
        document_files = entry.get('document_files', [])

        if not document_files:
            messagebox.showinfo("Информация", "Нет документов")
            return

        contractor_root = entry.get('contractor_root')
        root_path = entry.get('root_path', '')

        for doc_file in document_files:
            doc_found = False
            doc_paths_to_try = []

            # Различные варианты путей для поиска
            if os.path.isabs(doc_file):
                doc_paths_to_try.append(os.path.normpath(doc_file))

            # Относительный путь от папки с JSON
            doc_paths_to_try.append(os.path.normpath(os.path.join(root_path, doc_file)))

            # В корневой папке подрядчика
            if contractor_root:
                doc_paths_to_try.append(os.path.normpath(os.path.join(contractor_root, doc_file)))
                doc_paths_to_try.append(
                    os.path.normpath(os.path.join(contractor_root, "Документы", os.path.basename(doc_file))))

                # Рекурсивный поиск
                found_recursive = self.find_file_recursive(contractor_root, os.path.basename(doc_file))
                if found_recursive:
                    doc_paths_to_try.append(found_recursive)

            # Глобальный поиск
            found_global = self.find_file_recursive(self.current_directory, os.path.basename(doc_file))
            if found_global:
                doc_paths_to_try.append(found_global)

            # Пробуем открыть по найденным путям
            for doc_path in doc_paths_to_try:
                normalized_path = os.path.normpath(doc_path)
                if os.path.exists(normalized_path):
                    self.open_file(normalized_path)
                    doc_found = True
                    break

            if not doc_found:
                messagebox.showerror("Ошибка", f"Файл не найден: {os.path.basename(doc_file)}")

    def find_file_recursive(self, start_path, filename):
        """Рекурсивный поиск файла"""
        if not start_path or not os.path.exists(start_path):
            return None

        for root, dirs, files in os.walk(start_path):
            if filename in files:
                return os.path.join(root, filename)
        return None

    def open_file(self, file_path):
        try:
            path = os.path.normpath(file_path)
            if platform.system() == 'Darwin':
                subprocess.call(('open', path))
            elif platform.system() == 'Windows':
                os.startfile(path)
            else:
                subprocess.call(('xdg-open', path))
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл: {file_path}\n{str(e)}")

    def export_to_excel(self):
        """Экспорт данных в Excel"""
        if not self.data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return

        # Выбираем файл для сохранения
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"journal_incoming_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not filename:
            return

        try:
            # Создаем рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Журнал входного контроля"

            # Сортируем данные по дате для экспорта
            sorted_data = sorted(self.data, key=lambda x: self.parse_date(x.get('date', '')))

            # Заголовки столбцов в новом порядке
            headers = [
                "№ п/п",  # 1
                "Дата доставки",  # 2
                "Наименование деталей, материалов, изделий, конструкций, оборудования",  # 3
                "Кол-во",  # 4
                "Ед. изм.",  # 5
                "Поставщик",  # 6
                "Наименование и номер документа изготовителя",  # 7
                "Результат проверки сопроводительных документов производителя и визуального осмотра на соответствие требованиям утвержденной проектной документации и соответствующим документам по стандартизации",
                # 8
                "Решение о необходимости проведения лабораторного контроля",  # 9
                "Результат лабораторного контроля",  # 10
                # Остальные столбцы в произвольном порядке
                "Исполнитель",
                "Подрядчик",
                "Количество файлов",
                "Создано"
            ]

            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Записываем данные
            current_row = 2
            material_number = 0  # Счетчик для нумерации материалов

            for entry in sorted_data:
                date_value = entry.get('date', '')
                name_value = entry.get('name', '')

                # Увеличиваем номер материала
                material_number += 1

                # Формируем количество
                quantity_value = entry.get('quantity', '')
                if isinstance(quantity_value, (int, float)):
                    quantity_value = quantity_value

                unit_value = entry.get('quantity_unit', '')
                supplier_value = entry.get('supplier', '')
                document_value = entry.get('document', '')
                doc_check_value = entry.get('document_check_result', '')
                lab_control_value = "Да" if entry.get('lab_control_needed', False) else "Нет"
                lab_result_value = entry.get('lab_control_result', '')
                executor_value = entry.get('filled_by', '')
                contractor_value = entry.get('contractor', '')
                files_count = len(entry.get('document_files', []))
                created_value = entry.get('created_at', '')

                # Дата показывается для каждого материала (убрана группировка)
                display_date = date_value

                row_data = [
                    material_number,  # 1. № п/п
                    display_date,  # 2. Дата доставки (теперь для каждого материала)
                    name_value,  # 3. Наименование деталей, материалов...
                    quantity_value,  # 4. Кол-во
                    unit_value,  # 5. Ед. изм.
                    supplier_value,  # 6. Поставщик
                    document_value,  # 7. Наименование и номер документа изготовителя
                    doc_check_value,  # 8. Результат проверки сопроводительных документов...
                    lab_control_value,  # 9. Решение о необходимости проведения лабораторного контроля
                    lab_result_value,  # 10. Результат лабораторного контроля
                    # Остальные столбцы
                    executor_value,
                    contractor_value,
                    files_count if files_count > 0 else "",
                    created_value
                ]

                # Записываем строку данных
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = border

                    # Выравнивание для разных типов столбцов
                    if col in [1, 4, 13]:  # № п/п, Кол-во, Количество файлов
                        cell.alignment = Alignment(horizontal='center')
                    elif col == 2:  # Дата
                        cell.alignment = Alignment(horizontal='center')
                    elif col in [3, 6, 7, 8, 9, 10, 11, 12]:  # Текстовые поля
                        cell.alignment = Alignment(horizontal='left', wrap_text=True)

                current_row += 1

            # Автоподбор ширины столбцов
            column_widths = [
                8,  # № п/п
                12,  # Дата доставки
                40,  # Наименование деталей, материалов...
                10,  # Кол-во
                10,  # Ед. изм.
                20,  # Поставщик
                25,  # Наименование и номер документа изготовителя
                50,  # Результат проверки сопроводительных документов...
                15,  # Решение о необходимости проведения лабораторного контроля
                20,  # Результат лабораторного контроля
                20,  # Исполнитель
                20,  # Подрядчик
                12,  # Количество файлов
                20  # Создано
            ]

            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

            # Устанавливаем высоту строки заголовков для лучшего отображения
            ws.row_dimensions[1].height = 60

            # Добавляем автофильтр
            ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{current_row - 1}"

            # Замораживаем первую строку
            ws.freeze_panes = "A2"

            # Сохраняем файл
            wb.save(filename)
            messagebox.showinfo("Успех", f"Данные экспортированы в файл:\n{filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при экспорте в Excel: {str(e)}")

    def parse_date(self, date_str):
        """Парсинг даты для сортировки"""
        if not date_str:
            return datetime.min

        # Попробуем различные форматы дат
        date_formats = [
            "%Y-%m-%d",
            "%d.%m.%Y",
            "%d/%m/%Y",
            "%Y-%m-%d %H:%M:%S",
            "%d.%m.%Y %H:%M:%S"
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        # Если не удалось распарсить, возвращаем минимальную дату
        return datetime.min

    def save_data(self):
        if not self.data:
            messagebox.showwarning("Предупреждение", "Нет данных для сохранения")
            return

        # Завершаем текущее редактирование перед сохранением
        self.finish_edit()

        files_data = {}
        for entry in self.data:
            file_path = entry['file_path']
            if file_path not in files_data:
                files_data[file_path] = []
            clean_entry = {k: v for k, v in entry.items()
                           if k not in ['contractor', 'file_path', 'root_path', 'contractor_root']}
            files_data[file_path].append(clean_entry)

        saved_count = 0
        for file_path, entries in files_data.items():
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump({'entries': entries}, f, ensure_ascii=False, indent=2)
                saved_count += 1
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при сохранении {file_path}: {str(e)}")

        messagebox.showinfo("Успех", f"Сохранено {saved_count} файлов")


def main():
    root = tk.Tk()
    app = IncomingJournalEditor(root)
    root.mainloop()


if __name__ == "__main__":
    main()
