import json
import re
import sys
import threading
import queue
from pathlib import Path
from shutil import copy2
from typing import Dict, Any, List

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from openpyxl import load_workbook
from docxtpl import DocxTemplate

DATA_JSON = Path("data.json")
EXPORT_DIR = Path("Выгрузка")  # папка с шаблонами для подстановки
OUTPUT_DIR = Path("Вывод")  # папка для результатов
WORD_EXT = {".docx"}
EXCEL_EXT = {".xlsx"}

# Очередь событий для GUI-лога и прогресса
event_q: "queue.Queue[dict]" = queue.Queue()


def gui_log(msg: str):
    event_q.put({"type": "log", "msg": msg})


def gui_progress(current: int = None, total: int = None, step: int = None):
    """
    Обновление прогресса.
    - можно передать total один раз
    - либо step=1 для инкремента
    - либо current для явной установки
    """
    payload = {"type": "progress"}
    if total is not None:
        payload["total"] = total
    if current is not None:
        payload["current"] = current
    if step is not None:
        payload["step"] = step
    event_q.put(payload)


def load_mapping(data_json: Path) -> Dict[str, Any]:
    if not data_json.exists():
        raise FileNotFoundError(f"Не найден {data_json}")

    with data_json.open("r", encoding="utf-8") as f:
        items = json.load(f)

    mapping = {}
    for it in items:
        name = it.get("data_name")
        val = it.get("extracted_value")
        status = it.get("status")
        if not name:
            continue
        if status != "found":
            continue
        if val in (None, "null", ""):
            continue
        mapping[name] = val
    return mapping


def make_safe_key(s: str) -> str:
    safe = re.sub(r"[^\w]+", "_", s, flags=re.UNICODE).strip("_")
    return f"k_{safe}" if safe and not safe.isalpha() else (safe if safe else "k_key")


def build_context(mapping: Dict[str, Any]) -> Dict[str, Any]:
    ctx = {}
    for name, val in mapping.items():
        safe = make_safe_key(name)
        ctx[safe] = val
        if re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
            ctx[name] = val
    return ctx


def ensure_output_copy(src: Path) -> Path:
    rel = src.relative_to(EXPORT_DIR)
    dst = OUTPUT_DIR / rel
    dst.parent.mkdir(parents=True, exist_ok=True)
    copy2(src, dst)
    return dst


def process_word_file(path: Path, context: Dict[str, Any]) -> None:
    try:
        doc = DocxTemplate(str(path))
        doc.render(context)
        doc.save(str(path))
        gui_log(f"[WORD] Обновлён: {path}")
    except Exception as e:
        gui_log(f"[WORD] Ошибка {path}: {e}")


def process_excel_file(path: Path, mapping: Dict[str, Any]) -> None:
    try:
        wb = load_workbook(filename=str(path), data_only=False)
        changed = False
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("!!!"):
                        key = val[3:].strip()
                        if key in mapping:
                            cell.value = mapping[key]
                            changed = True
        if changed:
            wb.save(str(path))
            gui_log(f"[XLSX] Обновлён: {path}")
        else:
            gui_log(f"[XLSX] Без изменений: {path}")
    except Exception as e:
        gui_log(f"[XLSX] Ошибка {path}: {e}")


def find_files(base: Path) -> List[Path]:
    if not base.exists():
        gui_log(f"Папка {base} не найдена.")
        return []
    files = []
    for p in base.rglob("*"):
        if p.is_file() and p.suffix.lower() in (WORD_EXT | EXCEL_EXT):
            files.append(p)
    return files


def worker_run(data_json: Path, export_dir: Path, output_dir: Path):
    try:
        mapping = load_mapping(data_json)
    except Exception as e:
        gui_log(f"Не удалось загрузить {data_json}: {e}")
        event_q.put({"type": "done", "ok": False})
        return

    if not mapping:
        gui_log("В data.json нет валидных значений для подстановки (status!='found' или пустые).")

    context = build_context(mapping)

    files = find_files(export_dir)
    if not files:
        gui_log(f"В {export_dir} файлов Word/Excel не найдено.")
        event_q.put({"type": "done", "ok": False})
        return

    output_dir.mkdir(parents=True, exist_ok=True)

    total = len(files)
    gui_progress(total=total, current=0)
    cur = 0

    for src in files:
        try:
            dst = ensure_output_copy(src)
            ext = dst.suffix.lower()
            if ext in WORD_EXT:
                process_word_file(dst, context)
            elif ext in EXCEL_EXT:
                process_excel_file(dst, mapping)
        except Exception as e:
            gui_log(f"[COPY/PROC] Ошибка {src}: {e}")
        finally:
            cur += 1
            gui_progress(current=cur)

    gui_log(f"Готово. Результаты в: {output_dir.resolve()}")
    event_q.put({"type": "done", "ok": True})


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Выгрузка в шаблоны (Word/Excel)")
        self.geometry("820x520")
        self.minsize(720, 480)

        self.create_widgets()
        self.is_running = False

        # Периодический поллинг очереди
        self.after(100, self.poll_queue)

    def create_widgets(self):
        frm_top = ttk.Frame(self)
        frm_top.pack(fill="x", padx=10, pady=10)

        # Пути (с возможностью поменять)
        self.var_data = tk.StringVar(value=str(DATA_JSON))
        self.var_export = tk.StringVar(value=str(EXPORT_DIR))
        self.var_output = tk.StringVar(value=str(OUTPUT_DIR))

        def row_with_browse(parent, label, var, is_dir=True):
            row = ttk.Frame(parent)
            row.pack(fill="x", pady=4)
            ttk.Label(row, text=label, width=22).pack(side="left")
            entry = ttk.Entry(row, textvariable=var)
            entry.pack(side="left", fill="x", expand=True, padx=(5, 5))

            def browse():
                if is_dir:
                    path = filedialog.askdirectory(initialdir=".")
                else:
                    path = filedialog.askopenfilename(initialdir=".", filetypes=[("JSON", "*.json"), ("All", "*.*")])
                if path:
                    var.set(path)

            ttk.Button(row, text="Обзор...", command=browse).pack(side="right")
            return entry

        row_with_browse(frm_top, "Файл data.json:", self.var_data, is_dir=False)
        row_with_browse(frm_top, "Папка Выгрузка:", self.var_export, is_dir=True)
        row_with_browse(frm_top, "Папка Вывод:", self.var_output, is_dir=True)

        frm_prog = ttk.Frame(self)
        frm_prog.pack(fill="x", padx=10, pady=(0, 10))
        self.prog = ttk.Progressbar(frm_prog, orient="horizontal", mode="determinate", maximum=100)
        self.prog.pack(fill="x", expand=True)
        self.lbl_prog = ttk.Label(frm_prog, text="Ожидание...")
        self.lbl_prog.pack(anchor="w", pady=(4, 0))

        frm_btn = ttk.Frame(self)
        frm_btn.pack(fill="x", padx=10, pady=(0, 10))
        self.btn_start = ttk.Button(frm_btn, text="Старт", command=self.on_start)
        self.btn_start.pack(side="left")
        self.btn_stop = ttk.Button(frm_btn, text="Остановить", command=self.on_stop, state="disabled")
        self.btn_stop.pack(side="left", padx=(8, 0))

        frm_log = ttk.Frame(self)
        frm_log.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.txt = tk.Text(frm_log, wrap="word", state="disabled")
        self.txt.pack(side="left", fill="both", expand=True)
        scroll = ttk.Scrollbar(frm_log, command=self.txt.yview)
        scroll.pack(side="right", fill="y")
        self.txt.configure(yscrollcommand=scroll.set)

    def append_log(self, text: str):
        self.txt.configure(state="normal")
        self.txt.insert("end", text + "\n")
        self.txt.see("end")
        self.txt.configure(state="disabled")

    def set_progress(self, current: int, total: int):
        pct = 0 if total == 0 else int(current * 100 / total)
        self.prog["value"] = pct
        self.lbl_prog.config(text=f"Прогресс: {current}/{total} ({pct}%)")

    def poll_queue(self):
        try:
            while True:
                item = event_q.get_nowait()
                if item["type"] == "log":
                    self.append_log(item["msg"])
                elif item["type"] == "progress":
                    # Обновляем текущие и/или total
                    if "total" in item:
                        self.total = item["total"]
                        self.current = 0
                        self.set_progress(0, self.total)
                    if "current" in item:
                        self.current = item["current"]
                        self.set_progress(self.current, getattr(self, "total", 0))
                    if "step" in item:
                        self.current = getattr(self, "current", 0) + item["step"]
                        self.set_progress(self.current, getattr(self, "total", 0))
                elif item["type"] == "done":
                    self.is_running = False
                    self.btn_start.configure(state="normal")
                    self.btn_stop.configure(state="disabled")
                    ok = item.get("ok", True)
                    if ok:
                        self.append_log("Завершено успешно.")
                    else:
                        self.append_log("Завершено с ошибками или без результатов.")
        except queue.Empty:
            pass
        finally:
            self.after(100, self.poll_queue)

    def on_start(self):
        if self.is_running:
            return
        # Проверка путей
        data_p = Path(self.var_data.get()).expanduser()
        export_p = Path(self.var_export.get()).expanduser()
        output_p = Path(self.var_output.get()).expanduser()

        if not data_p.exists():
            messagebox.showerror("Ошибка", f"Не найден файл: {data_p}")
            return
        if not export_p.exists():
            messagebox.showerror("Ошибка", f"Не найдена папка: {export_p}")
            return
        output_p.mkdir(parents=True, exist_ok=True)

        # Очистка прогресса/лога
        self.txt.configure(state="normal")
        self.txt.delete("1.0", "end")
        self.txt.configure(state="disabled")
        self.set_progress(0, 0)
        self.append_log("Запуск...")

        # Запуск воркера
        self.is_running = True
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        t = threading.Thread(target=worker_run, args=(data_p, export_p, output_p), daemon=True)
        t.start()

    def on_stop(self):
        # Упрощённый стоп: сообщаем пользователю.
        # Без явной поддержки отмены в worker_run корректно прервать процесс нельзя.
        messagebox.showinfo("Остановить", "Мгновенная остановка не поддерживается. Закройте окно для прерывания.")
        # Можно реализовать флаг отмены и проверять его в цикле обработки файлов.


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
