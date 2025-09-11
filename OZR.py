import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import subprocess
import platform
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ProductionJournalEditor:
    def __init__(self, root):
        self.root = root
        self.root.title("Редактор журналов производства работ")
        self.root.geometry("1400x700")

        self.data = []
        self.current_directory = ""

        # Для inline редактирования
        self.edit_item = None
        self.edit_column = None
        self.entry_widget = None

        # Для сортировки
        self.sort_column = None
        self.sort_reverse = False

        self.create_widgets()

    def create_widgets(self):
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Button(top_frame, text="Выбрать директорию",
                   command=self.select_directory).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Обновить данные",
                   command=self.load_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Сохранить изменения",
                   command=self.save_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Экспорт в Excel",
                   command=self.export_to_excel).pack(side=tk.LEFT, padx=5)

        self.info_label = ttk.Label(top_frame, text="Выберите директорию для начала работы")
        self.info_label.pack(side=tk.LEFT, padx=20)

        # Инструкция по использованию
        instruction_label = ttk.Label(top_frame,
                                      text="Двойной клик - редактирование | Клик по заголовку - сортировка | Ctrl+Click - открыть фото")
        instruction_label.pack(side=tk.RIGHT, padx=10)

        self.create_treeview()

    def create_treeview(self):
        tree_frame = ttk.Frame(self.root)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Столбцы таблицы
        self.columns = ("Подрядчик", "Дата", "Наименование работ", "Объем", "Ед. изм.",
                        "Фото", "Исполнитель", "Создано")

        # Соответствие столбцов ключам в данных
        self.column_keys = {
            "Подрядчик": "contractor",
            "Дата": "date",
            "Наименование работ": "name",
            "Объем": "volume",
            "Ед. изм.": "volume_unit",
            "Фото": "photos",
            "Исполнитель": "filled_by",
            "Создано": "created_at"
        }

        # Редактируемые столбцы
        self.editable_columns = {"Дата", "Наименование работ", "Объем", "Ед. изм.", "Исполнитель"}

        self.tree = ttk.Treeview(tree_frame, columns=self.columns, show='headings', height=20)

        # Настройка заголовков и ширины столбцов
        column_widths = [120, 100, 250, 80, 80, 80, 150, 150]
        for i, col in enumerate(self.columns):
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_by_column(c))
            self.tree.column(col, width=column_widths[i])

        # Скроллбары
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Размещение элементов
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Привязка событий
        self.tree.bind('<Double-1>', self.on_double_click)
        self.tree.bind('<Control-1>', self.on_ctrl_click)
        self.tree.bind('<Button-1>', self.on_single_click)

    def select_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.current_directory = directory
            self.info_label.config(text=f"Директория: {directory}")
            self.load_data()

    def get_contractor_name(self, path):
        """Определение имени подрядчика из пути"""
        path_parts = os.path.normpath(path).split(os.sep)
        if "Журнал производства работ" in path_parts:
            idx = path_parts.index("Журнал производства работ")
            if idx > 0:
                return path_parts[idx - 1]
        return os.path.basename(os.path.dirname(path)) or "Неизвестный подрядчик"

    def find_contractor_root(self, contractor_name):
        """Поиск корневой папки подрядчика в основной директории"""
        if not self.current_directory:
            return None

        for root, dirs, files in os.walk(self.current_directory):
            if os.path.basename(root) == contractor_name:
                return root
        return None

    def load_data(self):
        if not self.current_directory:
            messagebox.showwarning("Предупреждение", "Сначала выберите директорию")
            return

        self.data = []

        for root, dirs, files in os.walk(self.current_directory):
            if "journal_production.json" in files:
                contractor = self.get_contractor_name(root)
                file_path = os.path.join(root, "journal_production.json")

                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        journal_data = json.load(f)

                    for entry in journal_data.get('entries', []):
                        entry['contractor'] = contractor
                        entry['file_path'] = file_path
                        entry['root_path'] = root
                        entry['contractor_root'] = self.find_contractor_root(contractor)
                        self.data.append(entry)

                except Exception as e:
                    messagebox.showerror("Ошибка", f"Ошибка при загрузке {file_path}: {str(e)}")

        self.update_treeview()
        self.info_label.config(text=f"Загружено записей: {len(self.data)}")

    def update_treeview(self):
        # Сохраняем текущий выбор
        selected = self.tree.selection()

        # Очищаем таблицу
        self.tree.delete(*self.tree.get_children())

        # Заполняем таблицы данными
        for i, entry in enumerate(self.data):
            photo_text = f"{len(entry.get('photos', []))} фото" if entry.get('photos') else "Нет"
            values = (
                entry.get('contractor', ''),
                entry.get('date', ''),
                entry.get('name', ''),
                entry.get('volume', ''),
                entry.get('volume_unit', ''),
                photo_text,
                entry.get('filled_by', ''),
                entry.get('created_at', '')
            )
            self.tree.insert('', 'end', iid=str(i), values=values)

        # Восстанавливаем выбор если возможно
        if selected and selected[0] in [self.tree.get_children()[i] for i in range(len(self.data))]:
            self.tree.selection_set(selected[0])

    def sort_by_column(self, column):
        """Сортировка по столбцу"""
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_reverse = False

        self.sort_column = column

        # Получаем ключ для сортировки
        sort_key = self.column_keys.get(column, column.lower())

        def sort_function(entry):
            value = entry.get(sort_key, '')
            if sort_key == 'volume':
                try:
                    return float(value) if value else 0
                except (ValueError, TypeError):
                    return 0
            elif sort_key == 'photos':
                return len(entry.get('photos', []))
            return str(value).lower()

        # Сортируем данные
        self.data.sort(key=sort_function, reverse=self.sort_reverse)

        # Обновляем отображение
        self.update_treeview()

        # Обновляем заголовок с индикатором сортировки
        for col in self.columns:
            if col == column:
                direction = " ↓" if self.sort_reverse else " ↑"
                self.tree.heading(col, text=col + direction)
            else:
                self.tree.heading(col, text=col)

    def export_to_excel(self):
        """Экспорт данных в Excel"""
        if not self.data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return

        # Выбираем файл для сохранения
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"journal_production_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            # Изменено с initialname на initialfile
        )

        if not filename:
            return

        try:
            # Создаем рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Журнал производства работ"

            # Сортируем данные по дате для экспорта
            sorted_data = sorted(self.data, key=lambda x: self.parse_date(x.get('date', '')))

            # Заголовки столбцов
            headers = [
                "Дата", "Наименование работ", "Объем", "Единица измерения",
                "Исполнитель", "Подрядчик", "Количество фото", "Создано"
            ]

            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Записываем данные
            current_row = 2
            previous_date = None

            for entry in sorted_data:
                date_value = entry.get('date', '')
                name_value = entry.get('name', '')
                volume_value = entry.get('volume', '')
                unit_value = entry.get('volume_unit', '')
                executor_value = entry.get('filled_by', '')
                contractor_value = entry.get('contractor', '')
                photos_count = len(entry.get('photos', []))
                created_value = entry.get('created_at', '')

                # Дата (показываем только если отличается от предыдущей)
                display_date = date_value if date_value != previous_date else ""

                row_data = [
                    display_date,
                    name_value,
                    volume_value,
                    unit_value,
                    executor_value,
                    contractor_value,
                    photos_count if photos_count > 0 else "",
                    created_value
                ]

                # Записываем строку данных
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = border

                    # Выравнивание для числовых столбцов
                    if col in [3, 7]:  # Объем и Количество фото
                        cell.alignment = Alignment(horizontal='right')
                    elif col == 1:  # Дата
                        cell.alignment = Alignment(horizontal='center')

                previous_date = date_value
                current_row += 1

            # Автоподбор ширины столбцов
            column_widths = [12, 35, 10, 15, 20, 20, 12, 20]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

            # Добавляем автофильтр
            ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{current_row - 1}"

            # Замораживаем первую строку
            ws.freeze_panes = "A2"

            # Сохраняем файл
            wb.save(filename)
            messagebox.showinfo("Успех", f"Данные экспортированы в файл:\n{filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при экспорте в Excel: {str(e)}")

    def parse_date(self, date_str):
        """Парсинг даты для сортировки"""
        if not date_str:
            return datetime.min

        # Попробуем различные форматы дат
        date_formats = [
            "%Y-%m-%d",
            "%d.%m.%Y",
            "%d/%m/%Y",
            "%Y-%m-%d %H:%M:%S",
            "%d.%m.%Y %H:%M:%S"
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        # Если не удалось распарсить, возвращаем минимальную дату
        return datetime.min

    def on_single_click(self, event):
        """Обработка одиночного клика - завершение редактирования"""
        self.finish_edit()

    def on_double_click(self, event):
        """Обработка двойного клика - начало редактирования"""
        self.finish_edit()  # Завершаем предыдущее редактирование

        item = self.tree.selection()[0] if self.tree.selection() else None
        if not item:
            return

        column = self.tree.identify_column(event.x)
        if not column:
            return

        column_index = int(column.replace('#', '')) - 1
        column_name = self.columns[column_index]

        # Проверяем, можно ли редактировать этот столбец
        if column_name not in self.editable_columns:
            return

        self.start_edit(item, column_name, column_index)

    def on_ctrl_click(self, event):
        """Обработка Ctrl+Click - открытие фото"""
        item = self.tree.selection()[0] if self.tree.selection() else None
        if item:
            self.open_photos_for_item(item)

    def start_edit(self, item, column_name, column_index):
        """Начало редактирования ячейки"""
        self.edit_item = item
        self.edit_column = column_name

        # Получаем координаты и размеры ячейки
        bbox = self.tree.bbox(item, column_index)
        if not bbox:
            return

        x, y, width, height = bbox

        # Получаем текущее значение
        current_value = self.tree.item(item, 'values')[column_index]

        # Создаем Entry для редактирования
        self.entry_widget = tk.Entry(self.tree)
        self.entry_widget.place(x=x, y=y, width=width, height=height)
        self.entry_widget.insert(0, current_value)
        self.entry_widget.select_range(0, tk.END)
        self.entry_widget.focus()

        # Привязываем события
        self.entry_widget.bind('<Return>', self.finish_edit)
        self.entry_widget.bind('<Escape>', self.cancel_edit)
        self.entry_widget.bind('<FocusOut>', self.finish_edit)

    def finish_edit(self, event=None):
        """Завершение редактирования"""
        if not self.entry_widget or not self.edit_item:
            return

        # Получаем новое значение
        new_value = self.entry_widget.get()

        # Обновляем данные
        item_index = int(self.edit_item)
        column_key = self.column_keys[self.edit_column]

        # Специальная обработка для числовых полей
        if column_key == 'volume':
            try:
                new_value = float(new_value) if new_value.strip() else None
            except ValueError:
                messagebox.showerror("Ошибка", "Объем должен быть числом")
                return

        # Обновляем данные
        self.data[item_index][column_key] = new_value

        # Обновляем время изменения
        self.data[item_index]['created_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Убираем Entry
        self.entry_widget.destroy()
        self.entry_widget = None
        self.edit_item = None
        self.edit_column = None

        # Обновляем отображение
        self.update_treeview()

    def cancel_edit(self, event=None):
        """Отмена редактирования"""
        if self.entry_widget:
            self.entry_widget.destroy()
            self.entry_widget = None
            self.edit_item = None
            self.edit_column = None

    def open_photos_for_item(self, item):
        """Открытие фото для конкретного элемента"""
        item_index = int(item)
        entry = self.data[item_index]
        photos = entry.get('photos', [])

        if not photos:
            messagebox.showinfo("Информация", "Нет фотографий")
            return

        contractor_name = entry.get('contractor', '')
        contractor_root = entry.get('contractor_root')

        for photo in photos:
            photo_found = False
            photo_paths_to_try = []

            # Различные варианты путей для поиска
            if os.path.isabs(photo):
                photo_paths_to_try.append(os.path.normpath(photo))

            if contractor_root:
                photo_paths_to_try.append(os.path.join(contractor_root, "Фотофиксация", os.path.basename(photo)))
                photo_paths_to_try.append(os.path.join(contractor_root, "Фотофиксация", photo))

                found_recursive = self.find_file_recursive(contractor_root, os.path.basename(photo))
                if found_recursive:
                    photo_paths_to_try.append(found_recursive)

            found_global = self.find_file_recursive(self.current_directory, os.path.basename(photo))
            if found_global:
                photo_paths_to_try.append(found_global)

            # Пробуем открыть по найденным путям
            for photo_path in photo_paths_to_try:
                normalized_path = os.path.normpath(photo_path)
                if os.path.exists(normalized_path):
                    self.open_file(normalized_path)
                    photo_found = True
                    break

            if not photo_found:
                messagebox.showerror("Ошибка", f"Файл не найден: {os.path.basename(photo)}")

    def find_file_recursive(self, start_path, filename):
        """Рекурсивный поиск файла"""
        if not start_path or not os.path.exists(start_path):
            return None

        for root, dirs, files in os.walk(start_path):
            if filename in files:
                return os.path.join(root, filename)
        return None

    def open_file(self, file_path):
        try:
            path = os.path.normpath(file_path)
            if platform.system() == 'Darwin':
                subprocess.call(('open', path))
            elif platform.system() == 'Windows':
                os.startfile(path)
            else:
                subprocess.call(('xdg-open', path))
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл: {file_path}\n{str(e)}")

    def save_data(self):
        if not self.data:
            messagebox.showwarning("Предупреждение", "Нет данных для сохранения")
            return

        # Завершаем текущее редактирование перед сохранением
        self.finish_edit()

        files_data = {}
        for entry in self.data:
            file_path = entry['file_path']
            if file_path not in files_data:
                files_data[file_path] = []
            clean_entry = {k: v for k, v in entry.items()
                           if k not in ['contractor', 'file_path', 'root_path', 'contractor_root']}
            files_data[file_path].append(clean_entry)

        saved_count = 0
        for file_path, entries in files_data.items():
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump({'entries': entries}, f, ensure_ascii=False, indent=2)
                saved_count += 1
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при сохранении {file_path}: {str(e)}")

        messagebox.showinfo("Успех", f"Сохранено {saved_count} файлов")


def main():
    root = tk.Tk()
    app = ProductionJournalEditor(root)
    root.mainloop()


if __name__ == "__main__":
    main()
